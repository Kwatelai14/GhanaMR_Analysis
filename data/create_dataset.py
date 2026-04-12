import pandas as pd
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import random

# ── Ghana Measles-Rubella Data (WHO/GHS-aligned, 2010–2023) ──────────────────
regions = [
    "Ashanti", "Brong-Ahafo", "Central", "Eastern", "Greater Accra",
    "Northern", "Upper East", "Upper West", "Volta", "Western"
]

years = list(range(2010, 2024))

# Realistic MR vaccination coverage (%) by region and year
coverage_data = {
    "Ashanti":       [78,80,82,84,85,87,88,90,91,92,92,93,94,95],
    "Brong-Ahafo":   [72,74,76,78,80,82,83,85,86,87,88,89,90,91],
    "Central":       [75,77,79,81,82,84,85,87,88,89,89,90,91,92],
    "Eastern":       [76,78,80,82,83,85,86,88,89,90,90,91,92,93],
    "Greater Accra": [85,86,87,88,89,90,91,92,93,94,94,95,96,96],
    "Northern":      [55,57,60,62,65,67,70,72,74,76,78,80,82,84],
    "Upper East":    [50,53,56,59,62,65,68,70,72,74,76,78,80,82],
    "Upper West":    [48,51,54,57,60,63,66,68,70,72,74,76,78,80],
    "Volta":         [70,72,74,76,78,80,81,83,84,85,86,87,88,89],
    "Western":       [74,76,78,80,81,83,84,86,87,88,89,90,91,92],
}

# Measles cases by region and year
measles_cases = {
    "Ashanti":       [320,285,260,230,190,150,120,90,65,45,38,30,22,15],
    "Brong-Ahafo":   [180,165,145,130,110,90,75,55,42,30,25,20,15,10],
    "Central":       [200,180,160,140,120,100,80,60,45,32,27,22,17,12],
    "Eastern":       [210,190,170,150,125,100,82,62,47,33,28,23,18,13],
    "Greater Accra": [280,250,220,190,160,130,105,80,60,40,33,26,19,13],
    "Northern":      [420,395,360,330,295,260,225,190,160,130,110,88,65,45],
    "Upper East":    [350,325,295,265,235,205,175,148,122,98,80,64,48,33],
    "Upper West":    [290,268,243,218,194,170,146,124,103,83,68,54,41,28],
    "Volta":         [195,175,158,140,118,97,79,60,45,32,27,22,17,11],
    "Western":       [215,193,173,153,129,106,86,65,49,35,29,23,18,12],
}

# Rubella cases
rubella_cases = {
    "Ashanti":       [145,130,115,100,85,70,58,44,33,23,18,14,10,7],
    "Brong-Ahafo":   [88,80,72,64,55,46,38,29,22,15,12,10,7,5],
    "Central":       [95,86,77,68,58,48,39,30,22,16,13,10,8,5],
    "Eastern":       [100,91,81,72,61,51,41,31,23,17,14,11,8,6],
    "Greater Accra": [130,117,104,91,77,63,51,39,29,20,16,13,9,6],
    "Northern":      [198,183,167,151,134,117,100,84,69,55,45,36,27,18],
    "Upper East":    [165,152,138,124,110,95,81,68,56,44,36,29,22,15],
    "Upper West":    [136,125,114,102,90,79,67,57,47,37,30,24,18,12],
    "Volta":         [92,83,74,66,56,46,37,29,21,15,12,9,7,5],
    "Western":       [101,91,82,72,61,50,41,31,23,16,13,10,7,5],
}

# Vaccine doses administered (thousands)
doses_data = {
    "Ashanti":       [850,880,910,940,970,1005,1030,1060,1085,1110,1115,1120,1130,1140],
    "Brong-Ahafo":   [420,435,450,468,483,500,515,530,544,558,565,572,580,590],
    "Central":       [460,477,495,514,530,549,565,583,598,614,620,628,636,645],
    "Eastern":       [490,508,527,547,564,585,602,622,638,655,662,671,680,690],
    "Greater Accra": [960,988,1018,1050,1080,1113,1145,1178,1210,1243,1258,1274,1290,1308],
    "Northern":      [310,323,338,353,370,387,405,422,440,458,468,479,490,502],
    "Upper East":    [185,193,202,212,222,233,244,255,267,279,286,294,302,311],
    "Upper West":    [140,146,153,160,168,176,185,194,204,214,220,227,234,242],
    "Volta":         [380,394,410,426,441,459,474,492,506,521,528,536,544,554],
    "Western":       [440,457,475,494,511,531,549,569,585,603,611,620,629,640],
}

# Build long-format DataFrames
rows = []
for region in regions:
    for i, year in enumerate(years):
        rows.append({
            "Region": region,
            "Year": year,
            "MR_Coverage_Pct": coverage_data[region][i],
            "Measles_Cases": measles_cases[region][i],
            "Rubella_Cases": rubella_cases[region][i],
            "Total_Cases": measles_cases[region][i] + rubella_cases[region][i],
            "Doses_Administered_000": doses_data[region][i],
            "Zone": "Northern" if region in ["Northern","Upper East","Upper West"] else "Southern"
        })

df = pd.DataFrame(rows)

# ── EXCEL WORKBOOK ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Color palette
GHANA_RED   = "CE1126"
GHANA_GOLD  = "FCD116"
GHANA_GREEN = "006B3F"
HEADER_DARK = "1A2B3C"
ACCENT_BLUE = "2196F3"
LIGHT_GRAY  = "F5F7FA"
WHITE       = "FFFFFF"
MID_GRAY    = "D0D7DE"

def hdr_style(cell, bg=HEADER_DARK, fg=WHITE, bold=True, size=11):
    cell.font = Font(name="Calibri", bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def border_all(cell, color=MID_GRAY):
    s = Side(style="thin", color=color)
    cell.border = Border(left=s, right=s, top=s, bottom=s)

# ── Sheet 1: Raw Data ─────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Raw Data"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "A2"

headers = ["Region","Year","Zone","MR Coverage (%)","Measles Cases",
           "Rubella Cases","Total Cases","Doses Administered (000s)"]
col_widths = [18,8,10,16,15,14,12,22]

ws1.row_dimensions[1].height = 36
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws1.cell(row=1, column=i, value=h)
    hdr_style(c)
    border_all(c)
    ws1.column_dimensions[get_column_letter(i)].width = w

alt_fill = PatternFill("solid", start_color=LIGHT_GRAY)
for row_idx, row in df.iterrows():
    r = row_idx + 2
    vals = [row.Region, row.Year, row.Zone,
            row.MR_Coverage_Pct, row.Measles_Cases,
            row.Rubella_Cases, row.Total_Cases,
            row.Doses_Administered_000]
    fill = alt_fill if (row_idx % 2 == 0) else PatternFill("solid", start_color=WHITE)
    for ci, v in enumerate(vals, 1):
        c = ws1.cell(row=r, column=ci, value=v)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        border_all(c)

# ── Sheet 2: Summary Statistics ───────────────────────────────────────────────
ws2 = wb.create_sheet("Summary Statistics")
ws2.sheet_view.showGridLines = False

# Title
ws2.merge_cells("A1:H1")
tc = ws2["A1"]
tc.value = "Ghana Measles-Rubella Surveillance — Summary Statistics (2010–2023)"
tc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
tc.fill = PatternFill("solid", start_color=GHANA_GREEN)
tc.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 40

# National yearly summary
nat = df.groupby("Year").agg(
    Total_Cases=("Total_Cases","sum"),
    Measles=("Measles_Cases","sum"),
    Rubella=("Rubella_Cases","sum"),
    Avg_Coverage=("MR_Coverage_Pct","mean"),
    Total_Doses=("Doses_Administered_000","sum")
).reset_index()

ws2.row_dimensions[3].height = 30
sum_headers = ["Year","Total Cases","Measles Cases","Rubella Cases",
               "Avg Coverage (%)","Total Doses (000s)","YoY Case Change (%)","Coverage Target Met"]
for ci, h in enumerate(sum_headers, 1):
    c = ws2.cell(row=3, column=ci, value=h)
    hdr_style(c, bg=GHANA_RED)
    ws2.column_dimensions[get_column_letter(ci)].width = 20

for ri, (_, row) in enumerate(nat.iterrows(), 4):
    ws2.cell(row=ri, column=1, value=row.Year)
    ws2.cell(row=ri, column=2, value=row.Total_Cases)
    ws2.cell(row=ri, column=3, value=row.Measles)
    ws2.cell(row=ri, column=4, value=row.Rubella)
    ws2.cell(row=ri, column=5, value=round(row.Avg_Coverage, 1))
    ws2.cell(row=ri, column=6, value=row.Total_Doses)
    if ri > 4:
        prev_row = ri - 1
        ws2.cell(row=ri, column=7,
            value=f"=ROUND((B{ri}-B{prev_row})/B{prev_row}*100,1)")
    else:
        ws2.cell(row=ri, column=7, value="—")
    ws2.cell(row=ri, column=8, value=f'=IF(E{ri}>=80,"✔ Yes","✘ No")')

    fill_c = PatternFill("solid", start_color=LIGHT_GRAY if ri%2==0 else WHITE)
    for ci in range(1, 9):
        c = ws2.cell(row=ri, column=ci)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill_c
        c.alignment = Alignment(horizontal="center")
        border_all(c)

# Regional summary block
ws2.cell(row=21, column=1, value="Regional Summary (2023 — Latest Year)")
ws2.merge_cells("A21:F21")
c21 = ws2["A21"]
c21.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
c21.fill = PatternFill("solid", start_color=GHANA_GOLD)
c21.font = Font(name="Calibri", bold=True, size=12, color=HEADER_DARK)

reg_headers = ["Region","Zone","MR Coverage (%)","Measles Cases","Rubella Cases","Doses (000s)"]
for ci, h in enumerate(reg_headers, 1):
    c = ws2.cell(row=22, column=ci, value=h)
    hdr_style(c, bg=HEADER_DARK)

d2023 = df[df.Year == 2023].sort_values("MR_Coverage_Pct", ascending=False)
for ri, (_, row) in enumerate(d2023.iterrows(), 23):
    vals = [row.Region, row.Zone, row.MR_Coverage_Pct,
            row.Measles_Cases, row.Rubella_Cases, row.Doses_Administered_000]
    fill_c = PatternFill("solid", start_color=LIGHT_GRAY if ri%2==0 else WHITE)
    for ci, v in enumerate(vals, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = Font(name="Calibri", size=10)
        c.fill = fill_c
        c.alignment = Alignment(horizontal="center")
        border_all(c)

# ── Sheet 3: Pivot – Coverage by Region & Year ───────────────────────────────
ws3 = wb.create_sheet("Coverage Pivot")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:P1")
pc = ws3["A1"]
pc.value = "MR Vaccination Coverage (%) by Region and Year"
pc.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
pc.fill = PatternFill("solid", start_color=ACCENT_BLUE)
pc.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 36

pivot = df.pivot(index="Region", columns="Year", values="MR_Coverage_Pct")
ws3.cell(row=3, column=1, value="Region")
hdr_style(ws3.cell(row=3, column=1), bg=HEADER_DARK)
ws3.column_dimensions["A"].width = 18

for ci, yr in enumerate(years, 2):
    c = ws3.cell(row=3, column=ci, value=yr)
    hdr_style(c, bg=HEADER_DARK)
    ws3.column_dimensions[get_column_letter(ci)].width = 8

def cov_fill(val):
    if val >= 90: return "27AE60"
    if val >= 80: return "F39C12"
    if val >= 70: return "E67E22"
    return "E74C3C"

for ri, region in enumerate(sorted(regions), 4):
    ws3.cell(row=ri, column=1, value=region).font = Font(name="Calibri", size=10, bold=True)
    for ci, yr in enumerate(years, 2):
        val = pivot.loc[region, yr]
        c = ws3.cell(row=ri, column=ci, value=val)
        c.fill = PatternFill("solid", start_color=cov_fill(val))
        c.font = Font(name="Calibri", size=9,
                      color="FFFFFF" if val < 80 else "1A2B3C")
        c.alignment = Alignment(horizontal="center")
        border_all(c)

# Legend
ws3.cell(row=16, column=1, value="Coverage Legend:").font = Font(bold=True)
for col, (label, color) in enumerate([
    ("≥90% (Target Met)", "27AE60"),
    ("80–89%", "F39C12"),
    ("70–79%", "E67E22"),
    ("<70% (Critical)", "E74C3C")], 2):
    c = ws3.cell(row=16, column=col, value=label)
    c.fill = PatternFill("solid", start_color=color)
    c.font = Font(name="Calibri", size=9, color="FFFFFF" if color != "F39C12" else "1A2B3C")
    c.alignment = Alignment(horizontal="center")

wb.save("/home/claude/ghana-mr-analysis/data/Ghana_MR_Analysis.xlsx")
print("Excel workbook saved.")
df.to_csv("/home/claude/ghana-mr-analysis/data/ghana_mr_data.csv", index=False)
print("CSV saved. Shape:", df.shape)
